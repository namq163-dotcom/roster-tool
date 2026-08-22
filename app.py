import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import numpy as np
from datetime import datetime
from io import BytesIO, StringIO
from openpyxl import load_workbook

# ==========================================
# CẤU HÌNH GIAO DIỆN STREAMLIT (UI)
# ==========================================
st.set_page_config(page_title="Global APIS Automation", page_icon="✈️", layout="wide")

# CSS TUỲ CHỈNH ĐỂ GIAO DIỆN RỰC RỠ, BẮT MẮT VÀ CHUYÊN NGHIỆP HƠN
st.markdown("""
<style>
    .stButton > button {
        border-radius: 10px;
        font-weight: 600;
        transition: all 0.3s ease;
    }
    .stButton > button:hover {
        transform: translateY(-2px);
        box-shadow: 0 4px 12px rgba(0,0,0,0.15);
    }
</style>
""", unsafe_allow_html=True)

# KHUNG HEADER: GIAO DIỆN CHUẨN OCC HÀNG KHÔNG
header_html = """
<div style="display: flex; justify-content: space-between; align-items: center; padding: 14px 22px; background: linear-gradient(135deg, #f8f9fa, #e9ecef); border-radius: 10px; border: 1px solid #dcdfe6; font-family: 'Segoe UI', Roboto, Helvetica, Arial, sans-serif; margin-bottom: 20px; box-shadow: 0 3px 6px rgba(0,0,0,0.06);">
    <div style="display: flex; align-items: center;">
        <div style="background: linear-gradient(135deg, #0056b3, #00264d); padding: 10px 14px; border-radius: 8px; margin-right: 14px; box-shadow: 0 2px 5px rgba(0,0,0,0.2);">
            <span style="font-size: 22px;">✈️</span>
        </div>
        <div style="line-height: 1.3;">
            <div style="color: #00264d; font-size: 21px; font-weight: 800; letter-spacing: 0.5px; text-transform: uppercase;">CRYSTAL BAY AIRLINES</div>
            <div style="color: #0056b3; font-size: 11px; font-weight: 700; letter-spacing: 2.5px; margin-top: 2px;">APIS OPERATIONS CENTER</div>
        </div>
    </div>
    
    <div style="display: flex; gap: 12px; font-size: 11.5px; color: #2c3e50; background: #ffffff; padding: 8px 14px; border-radius: 8px; border: 1px solid #dcdfe6; box-shadow: inset 0 1px 3px rgba(0,0,0,0.03);">
        <div style="line-height: 1.6;">
            <div><b style="color: #c0392b;">🇻🇳 VN (Local):</b> <span id="time-vn" style="font-family: 'Consolas', monospace; font-size: 12px; font-weight: 700;"></span></div>
            <div><b style="color: #2980b9;">🌐 UTC:</b> <span id="time-utc" style="font-family: 'Consolas', monospace; font-size: 12px; font-weight: 700;"></span></div>
        </div>
        <div style="border-left: 1px solid #dcdfe6; padding-left: 10px; line-height: 1.6;">
            <div><b>🇰🇿 KZ - Kazakhstan:</b> <span id="time-kz" style="font-family: 'Consolas', monospace; font-size: 11.5px;"></span></div>
            <div><b>🇰🇬 KG - Kyrgyzstan:</b> <span id="time-kg" style="font-family: 'Consolas', monospace; font-size: 11.5px;"></span></div>
            <div><b>🇹🇯 TJ - Tajikistan:</b> <span id="time-tj" style="font-family: 'Consolas', monospace; font-size: 11.5px;"></span></div>
        </div>
        <div style="border-left: 1px solid #dcdfe6; padding-left: 10px; line-height: 1.6;">
            <div><b>🇷🇺 RU - Russia:</b> <span id="time-ru" style="font-family: 'Consolas', monospace; font-size: 11.5px;"></span></div>
            <div><b>🇵🇱 PL - Poland:</b> <span id="time-pl" style="font-family: 'Consolas', monospace; font-size: 11.5px;"></span></div>
        </div>
    </div>
</div>

<script>
    function updateTime() {
        const now = new Date();
        document.getElementById('time-utc').innerText = now.toLocaleTimeString('en-GB', {timeZone: 'UTC'});
        document.getElementById('time-vn').innerText = now.toLocaleTimeString('en-GB', {timeZone: 'Asia/Ho_Chi_Minh'});
        document.getElementById('time-kz').innerText = now.toLocaleTimeString('en-GB', {timeZone: 'Asia/Almaty'});
        document.getElementById('time-kg').innerText = now.toLocaleTimeString('en-GB', {timeZone: 'Asia/Bishkek'});
        document.getElementById('time-tj').innerText = now.toLocaleTimeString('en-GB', {timeZone: 'Asia/Dushanbe'});
        document.getElementById('time-ru').innerText = now.toLocaleTimeString('en-GB', {timeZone: 'Europe/Moscow'});
        document.getElementById('time-pl').innerText = now.toLocaleTimeString('en-GB', {timeZone: 'Europe/Warsaw'});
    }
    setInterval(updateTime, 1000);
    updateTime();
</script>
"""
components.html(header_html, height=110)

# ==========================================
# CÁC HÀM XỬ LÝ DỮ LIỆU
# ==========================================
def parse_date(date_str):
    if pd.isna(date_str) or not str(date_str).strip() or str(date_str).strip().lower() == 'nan': return ""
    date_str = str(date_str).strip()
    try:
        d = datetime.strptime(date_str, "%d/%m/%y")
        if d.year > 2050: d = d.replace(year=d.year - 100)
        return d.strftime("%d/%m/%Y")
    except: return date_str

def process_roster_data_vn(gd_file, template_file_path):
    content = gd_file.getvalue()
    df_gd = None
    try: df_gd = pd.read_excel(BytesIO(content))
    except: pass

    if df_gd is None or len(df_gd) == 0:
        for enc in ['utf-8', 'latin1', 'cp1258', 'utf-16']:
            try:
                dfs = pd.read_html(StringIO(content.decode(enc, errors='ignore')))
                if len(dfs) > 0: df_gd = dfs[0]; break
            except: continue

    if df_gd is None or len(df_gd) == 0: raise ValueError("Không thể đọc được dữ liệu trong file GD.")

    header_idx = None
    for idx, row in df_gd.iterrows():
        row_str = row.fillna("").astype(str).str.lower()
        if any('passport' in s for s in row_str) and any('name' in s for s in row_str):
            header_idx = idx; break
            
    if header_idx is None: raise ValueError("Không tìm thấy bảng danh sách tổ bay trong file GD.")
        
    header_row = df_gd.iloc[header_idx].fillna("").astype(str).str.lower()
    col_name = next((i for i, v in enumerate(header_row) if 'name' in v), None)
    col_passport = next((i for i, v in enumerate(header_row) if 'passport' in v and 'expiry' not in v), None)
    col_dob = next((i for i, v in enumerate(header_row) if 'birth' in v), None)
    col_gender = next((i for i, v in enumerate(header_row) if v == 'g'), None)
    col_nat = next((i for i, v in enumerate(header_row) if 'ntly' in v), None)
    col_expiry = next((i for i, v in enumerate(header_row) if 'expiry' in v), None)

    crew_data = []
    seen_passports = set()
    
    for idx in range(header_idx + 1, len(df_gd)):
        row = df_gd.iloc[idx]
        if 'declaration of health' in row.fillna("").astype(str).str.cat(sep=" ").lower(): break
            
        name_val = str(row.iloc[col_name]).strip() if col_name is not None and pd.notna(row.iloc[col_name]) else 'nan'
        passport_val = str(row.iloc[col_passport]).strip() if col_passport is not None and pd.notna(row.iloc[col_passport]) else 'nan'
        
        if name_val.lower() != 'nan' and passport_val.lower() != 'nan' and name_val != '':
            if passport_val in seen_passports: continue
            seen_passports.add(passport_val)
            name_parts = name_val.split()
            if len(name_parts) > 1 and len(name_parts[-1]) <= 3 and name_parts[-1].isupper(): name_parts = name_parts[:-1]
                
            family_name = name_parts[0] if name_parts else ""
            given_name = " ".join(name_parts[1:]) if len(name_parts) > 1 else ""
            nat = str(row.iloc[col_nat]).strip() if col_nat is not None and pd.notna(row.iloc[col_nat]) else ""
            if nat.upper() == 'VNM': nat = 'VN'
            gender = str(row.iloc[col_gender]).strip() if col_gender is not None and pd.notna(row.iloc[col_gender]) else ""
            
            crew_data.append([
                family_name, None, given_name, gender, nat, 
                parse_date(row.iloc[col_dob]) if col_dob is not None else "", 
                'P', passport_val, nat, 
                parse_date(row.iloc[col_expiry]) if col_expiry is not None else ""
            ])

    output = BytesIO()
    book = load_workbook(template_file_path)
    sheet = book.active
    
    for row in sheet.iter_rows(min_row=14, max_row=sheet.max_row, min_col=1, max_col=10):
        for cell in row: cell.value = None

    for r_idx, row_data in enumerate(crew_data, 14):
        for c_idx, value in enumerate(row_data, 1):
            sheet.cell(row=r_idx, column=c_idx, value=value)
            
    book.save(output)
    df_preview = pd.DataFrame(crew_data, columns=['Họ', 'Tên đệm', 'Tên', 'Giới tính', 'Quốc tịch', 'Ngày sinh', 'Loại giấy tờ', 'Số giấy tờ', 'Nơi cấp', 'Ngày hết hạn'])
    return output.getvalue(), df_preview

# ==========================================
# GIAO DIỆN CHỌN QUỐC GIA (Ô VUÔNG & CỜ THẬT)
# ==========================================
st.markdown("<p style='font-size: 16px; font-weight: bold; color: #1f2937;'>🌍 Vui lòng chọn Quốc gia đến để xuất APIS:</p>", unsafe_allow_html=True)

COUNTRY_CONFIG = {
    "Việt Nam (VNAPIS)": {"flag_url": "https://flagcdn.com/w80/vn.png", "name": "Việt Nam (VNAPIS)", "code": "VNAPIS", "template": "Template_VNAPIS.xlsx", "ready": True},
    "Kazakhstan": {"flag_url": "https://flagcdn.com/w80/kz.png", "name": "Kazakhstan", "code": "KZ", "template": "Template_Kazakhstan.xlsx", "ready": False},
    "Kyrgyzstan": {"flag_url": "https://flagcdn.com/w80/kg.png", "name": "Kyrgyzstan", "code": "KG", "template": "Template_Kyrgyzstan.xlsx", "ready": False},
    "Tajikistan": {"flag_url": "https://flagcdn.com/w80/tj.png", "name": "Tajikistan", "code": "TJ", "template": "Template_Tajikistan.xlsx", "ready": False},
    "Russia": {"flag_url": "https://flagcdn.com/w80/ru.png", "name": "Russia (Nga)", "code": "RU", "template": "Template_Russia.xlsx", "ready": False},
    "Poland": {"flag_url": "https://flagcdn.com/w80/pl.png", "name": "Poland (Ba Lan)", "code": "PL", "template": "Template_Poland.xlsx", "ready": False}
}

if 'selected_country' not in st.session_state:
    st.session_state.selected_country = "Việt Nam (VNAPIS)"

keys = list(COUNTRY_CONFIG.keys())

# Hiển thị 6 ô vuông chọn quốc gia trực quan thành 2 hàng (mỗi hàng 3 ô)
def render_country_grid(start_idx, end_idx):
    cols = st.columns(3)
    for i, idx in enumerate(range(start_idx, end_idx)):
        c_key = keys[idx]
        cfg = COUNTRY_CONFIG[c_key]
        is_selected = (st.session_state.selected_country == c_key)
        
        with cols[i]:
            # Tạo card màu sắc nổi bật hơn
            border_color = "#0056b3" if is_selected else "#d1d5db"
            bg_color = "#e6f0fa" if is_selected else "#ffffff"
            shadow = "0 4px 10px rgba(0,86,179,0.2)" if is_selected else "0 2px 4px rgba(0,0,0,0.05)"
            
            st.markdown(f"""
            <div style="display: flex; align-items: center; padding: 10px 14px; background-color: {bg_color}; border: 2px solid {border_color}; border-radius: 10px; box-shadow: {shadow}; margin-bottom: 8px;">
                <img src="{cfg['flag_url']}" width="38" style="border-radius: 4px; box-shadow: 0 1px 3px rgba(0,0,0,0.2); margin-right: 12px; object-fit: cover;">
                <div style="font-weight: 700; color: #1f2937; font-size: 14px;">{cfg['name']}</div>
            </div>
            """, unsafe_allow_html=True)
            
            btn_text = f"👉 Chọn {cfg['name'].split(' ')[0]}" if not is_selected else "✅ Đang chọn"
            if st.button(btn_text, key=f"btn_{c_key}", use_container_width=True):
                st.session_state.selected_country = c_key
                st.rerun()

render_country_grid(0, 3)
render_country_grid(3, 6)

st.markdown("<div style='margin-bottom: 15px;'></div>", unsafe_allow_html=True)

# Xử lý quốc gia được chọn
selected_country = st.session_state.selected_country
config = COUNTRY_CONFIG[selected_country]

if config["ready"]:
    uploaded_gd = st.file_uploader(f"Tải lên file GD (.xls, .xlsx) cho {config['name']}", type=["xls", "xlsx", "txt", "csv"])
    if uploaded_gd is not None:
        st.info(f"Đang xử lý dữ liệu cho {config['name']}...")
        try:
            if "Việt Nam" in selected_country:
                excel_data, preview_data = process_roster_data_vn(uploaded_gd, config["template"])
                
            st.success(f"✅ Đã xử lý thành công form APIS cho {config['name']}!")
            st.dataframe(preview_data) 
            st.download_button(
                label=f"⬇️ Tải form Excel hoàn chỉnh",
                data=excel_data,
                file_name=f"APIS_Crew_{config['code']}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except Exception as e:
            st.error(f"❌ Có lỗi xảy ra: {e}")
else:
    st.warning(f"🚧 Chức năng xuất APIS cho **{config['name']}** đang được xây dựng.")
    st.info("💡 **Hướng dẫn cho Admin:**\n1. Chuẩn bị file Excel mẫu của quốc gia này.\n2. Tải file mẫu lên hệ thống.\n3. Cung cấp quy tắc điền để lập trình viên hoàn thiện logic.")
